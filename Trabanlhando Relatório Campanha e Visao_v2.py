#!/usr/bin/env python
# coding: utf-8

# In[1]:


from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait
import pandas as pd
from bs4 import BeautifulSoup
import time
import datetime
import openpyxl
#para reproduzir em segundo plano
from selenium.webdriver.chrome.options import Options


# In[3]:


chrome_options = Options()
chrome_options.add_argument("-headless")
navegador = webdriver.Chrome()
#navegador = webdriver.Chrome(options = chrome_options) #- navegadro minimizado
data = datetime.datetime.now()


# In[4]:


url = 'http://10.10.220.66/Olos/Login.aspx'
campo_login = '//*[@id="UserTxt"]'
login = 'python_andre' #python_andre - andre_lucas
campo_senha = '//*[@id="Password"]'
senha = 'olos123'

navegador.get(url)
navegador.find_element(By.XPATH, campo_login).send_keys(login)      
navegador.find_element(By.XPATH, campo_senha).send_keys(senha + Keys.RETURN)


# In[5]:


#função para clicar em um path: 

def clica_forte(b,*c):
    clica_em = lambda a: navegador.find_element(By.XPATH, a).click()

    i=0
    while (i < 240):
        time.sleep(0.1)
        i += 1
        try:
            clica_em(b)
        except:
            continue
        else:
            print(f'{c}')
            break


# In[6]:


#Abre a aba de monitoramento:
monitoramento = '//*[@id="ctl00_TopMenu_Monitor"]'
clica_forte(monitoramento,'clicou con sucessso em monitoramenro!')

print('passou da aberturamonitoramento')


# In[7]:


#seleciona a área de relatórios
relatorios_1 = '//*[@id="ctl00_TopMenu_Reports"]'
relatorios_2 = '//*[@id="PageMenu_lblMenuLatReports"]'
#time.sleep(1)
clica_forte(relatorios_1)
clica_forte(relatorios_2)



# ## abre as abas  de Campanhas e Tabulação

# In[8]:


#seleciona a área de relatório visão da campanha:
visão_da_campanha = '//*[@id="PageMenu_menu1_labelMenuTitle_reports_view_campaign"]'
campanha = '//*[@id="PageMenu_menu1_submenu_reports_campaign"]'
clica_forte(visão_da_campanha)
clica_forte(campanha)


# In[9]:


#coloca as datas
campo_data_inicial = '//*[@id="PageContent_search1_StartDate"]'
campo_data_final = '//*[@id="PageContent_search1_EndDate"]'
data = f'{data.day:02}/{data.month:02}/{data.year}'

navegador.find_element(By.XPATH, campo_data_inicial).send_keys(data)
navegador.find_element(By.XPATH,campo_data_final).send_keys(data)


# In[10]:


#coloca o templete
campo_templete = '//*[@id="PageContent_search1_DDTemplate"]'
navegador.find_element(By.XPATH, campo_templete).send_keys('Olos Comparativo' + Keys.ARROW_UP)

#seleciona o formato atml e abre a nova aba com as informações de relatório:
formato_html = '//*[@id="PageContent_search1_btn_html"]'
clica_forte(formato_html)
time.sleep(5)
print('passou 05 segundos')


# In[11]:


#seleciona o relatório mailing x tabulação

visao_da_tabulacao = '//*[@id="PageMenu_menu1_labelMenuTitle_reports_view_disposition"]'
mailing_versus_x_tabulacao = '//*[@id="PageMenu_menu1_submenu_reports_dmailingd"]'

clica_forte(visao_da_tabulacao)
clica_forte(mailing_versus_x_tabulacao)

navegador.find_element(By.XPATH, campo_data_inicial).send_keys(data)
navegador.find_element(By.XPATH,campo_data_final).send_keys(data)
navegador.find_element(By.XPATH, campo_templete).send_keys('Discagem por mailing_2')
clica_forte(formato_html)
time.sleep(7)
print('passou 07 segundos')


# In[12]:


#funão para selecionar diferentes abas
navegador.current_window_handle
winds = navegador.window_handles

def find_window(url: str):
    for window in winds:
        navegador.switch_to.window(window)
        if url in navegador.current_url:
            break    


# ## trabalha os dados da Campanha

# In[13]:


find_window('CampaignReport')


# In[14]:


#colhe e instancia o patf da tabela campanha:
path_table_campanha =  '//*[@id="PageContent_tabela1_TabList"]'
element_campanha = navegador.find_element(By.XPATH, path_table_campanha)


# In[15]:


#Colhe o HTML de todas a tabelas:
html_content_table_campanha = element_campanha.get_attribute('outerHTML')


# In[16]:


#transforma o HTML em uma tabela

soup = BeautifulSoup(html_content_table_campanha, 'html.parser')
table = soup.find(name='table')


# In[17]:


df_full_table_campanha = pd.read_html(str(table))[0]


# In[20]:


coluna_1_campanha = df_full_table_campanha['CAMPANHA']
coluna_2_campanha = df_full_table_campanha['TEMPO LOGADO']
coluna_3_campanha = df_full_table_campanha['TEMPO FALANDO']
coluna_4_campanha = df_full_table_campanha['TEMPO DE TABULAÇÃO']
coluna_5_campanha = df_full_table_campanha['TEMPO PAUSA']
coluna_6_campanha = df_full_table_campanha['TEMPO OCIOSO']
coluna_7_campanha = df_full_table_campanha['TEMPO PRODUTIVO']
coluna_8_campanha = df_full_table_campanha['TEMPO DE CONVERSAÇÃO']
coluna_9_campanha = df_full_table_campanha['TEMPO OUTRA CAMPANHA']
coluna_10_campanha = df_full_table_campanha['TEMPO ESTADO CHAMADA MANUAL']
coluna_11_campanha = df_full_table_campanha['TEMPO CHAMADAS MANUAIS']
coluna_12_campanha = df_full_table_campanha['TEMPO MÉDIO DE ATENDIMENTO']
coluna_13_campanha = df_full_table_campanha['TEMPO MÉDIO DE TABULAÇÃO']
coluna_14_campanha = df_full_table_campanha['TEMPO MÉDIO OCIOSO']
coluna_15_campanha = df_full_table_campanha['TOTAL DE CHAMADAS']
coluna_16_campanha = df_full_table_campanha['CHAMADAS ATENDIDAS']
coluna_17_campanha = df_full_table_campanha['TOTAL DE CONTATOS']
coluna_18_campanha = df_full_table_campanha['CONTATO EFETIVO']
coluna_19_campanha = df_full_table_campanha['CHAMADAS MANUAIS']
coluna_20_campanha = df_full_table_campanha['% DE TEMPO FALADO']
coluna_21_campanha = df_full_table_campanha['% DE TEMPO TABULANDO']
coluna_22_campanha = df_full_table_campanha['% DE TEMPO EM PAUSA']
coluna_23_campanha = df_full_table_campanha['% DE TEMPO LIVRE']
coluna_24_campanha = df_full_table_campanha['% ATENDIDAS']
coluna_25_campanha = df_full_table_campanha['% CONTATOS']
coluna_26_campanha = df_full_table_campanha['% CONTATOS EFETIVOS / CHAMADAS ATENDIDAS']
coluna_27_campanha = df_full_table_campanha['% CONVERSÃO POR CONTATO EFETIVO']


colunas_campanha =[
    coluna_1_campanha,
    coluna_2_campanha,
    coluna_3_campanha,
    coluna_4_campanha,
    coluna_5_campanha,
    coluna_6_campanha,
    coluna_7_campanha,
    coluna_8_campanha,
    coluna_9_campanha,
    coluna_10_campanha,
    coluna_11_campanha,
    coluna_12_campanha,
    coluna_13_campanha,
    coluna_14_campanha,
    coluna_15_campanha,
    coluna_16_campanha,
    coluna_17_campanha,
    coluna_18_campanha,
    coluna_19_campanha,
    coluna_20_campanha,
    coluna_21_campanha,
    coluna_22_campanha,
    coluna_23_campanha,
    coluna_24_campanha,
    coluna_25_campanha,
    coluna_26_campanha,
    coluna_27_campanha
]


# ## Trabalha os dados da Visão:

# In[21]:


find_window('DispositionMailingDispositionReport.')


# In[22]:


#colhe e instancia o patf da tabela campanha:
path_table_visao =  '//*[@id="PageContent_tabela1_TabList"]'
element_visão = navegador.find_element(By.XPATH, path_table_visao)


# In[23]:


#Colhe o HTML de todas a visao:
html_content_table_visão = element_visão.get_attribute('outerHTML')


# In[24]:


#transforma o HTML em uma tabela

soup = BeautifulSoup(html_content_table_visão, 'html.parser')
table = soup.find(name='table')


# In[25]:


df_full_table_visao = pd.read_html(str(table))[0]


# In[43]:


df_full_table_visao['TOTAL DE CHAMADAS'] = df_full_table_visao['TOTAL DE CHAMADAS'].apply(lambda x: x.replace(".",""))
df_full_table_visao['TOTAL DE CHAMADAS'] = df_full_table_visao['TOTAL DE CHAMADAS'].astype(int)


# In[45]:


coluna_1_visao = df_full_table_visao['MAILING']
coluna_2_visao = df_full_table_visao['TABULAÇÃO']
coluna_3_visao = df_full_table_visao['TOTAL DE CHAMADAS']
coluna_4_visao = df_full_table_visao['CHAMADA MANUAL']


# In[46]:


coluna_tres_visao = df_full_table_visao['TOTAL DE CHAMADAS']


# In[43]:


##transformando a coluna 3 e 4 em INT, para qe o excel possa reconhecer seus valores como tal:

#fórmula para tranformar coluna STR em INT:

def transforma_coluna_em_int(coluna):
    lista7=[]
    for rows in coluna:
        try:
            rows = float(rows)
        except:
            rows = 0
        lista7.append(rows)
    return lista7

#executa a transformação:

coluna_3_visao = transforma_coluna_em_int(coluna_3_visao)
coluna_4_visao = transforma_coluna_em_int(coluna_4_visao)


# # começa edição

# In[44]:


coluna_tres_visao

for i in coluna_tres_visao:
    print(i)


# In[45]:


coluna_3_visao

for i in coluna_3_visao:
    print(i)


# In[25]:


##transformando a coluna 3 e 4 em INT, sem o ponto para qe o excel possa reconhecer seus valores como tal:

#fórmula para tranformar coluna STR em INT:

def transforma_coluna_em_int(coluna):
    lista7=[]
    for rows in coluna:
        try:
            rows = int(rows)
        except:
            rows = 0
        lista7.append(rows)
    return lista7

#executa a transformação:

coluna_3_visao = transforma_coluna_em_int(coluna_3_visao)


# In[26]:


coluna_3_visao = transforma_coluna_em_int(coluna_3_visao)


# In[27]:


coluna_3_visao


# In[122]:


a = str(18.000)


b = []
for i in coluna_3_visao:
    print(i)
    if i == '.':
        continue
    b.append(i)
print(b)



# In[97]:


for i in coluna_3_visao:
    print(i)


# # termina edição

# In[25]:


colunas_visao = [
    coluna_1_visao,
    coluna_2_visao,
    coluna_3_visao,
    coluna_4_visao
]


# ## Trabalha os dados do monitor:

# find_window('reports/disposition/DispositionMailingDisposition.aspx')

# #função para clicar em um path: 
# 
# paniga_inicial_1 = '//*[@id="linkStartPage"]'
# paniga_inicial_2 = '//*[@id="ctl00_labelStartPage"]'
# 
# clica_forte(paniga_inicial_1)
# print('passou da  pagina inicial 1')
# clica_forte(paniga_inicial_2)
# print('passou da  pagina inicial 2')
# 

# In[26]:


#monitoramento = '//*[@id="ctl00_TopMenu_Monitor"]'
#clica_forte(monitoramento)
#print('clicou em monitoramento')

time.sleep(10)
print('eperou 10 segundos')

# atualizando a funão para selecionar diferentes abas
navegador.current_window_handle
winds = navegador.window_handles

def find_window(url: str):
    for window in winds:
        navegador.switch_to.window(window)
        if url in navegador.current_url:
            break   
find_window('OlosMonitor/')
print('entrou em olos monitor um em outra aba!')


# sucesso em clicar! \
# clicou em pagina inicial \
# sucesso em clicar! \
# clicou em monitoramento \
# eperou 10 segundos \
# entrou em olos monitor um em outra aba!

# In[27]:


campanhas1 = '//*[@id="app"]/div/div[1]/div/div[1]/div/header/div[1]/div/a[1]/div/div/span'
campanhas2 = '//*[@id="app"]/div/div[1]/div/div[1]/div/header/div[1]/div/a[1]'
campanhas3 = '//*[@id="app"]/div/aside/div/div[2]/div/div/div[1]/div/div[2]'
campanhas4 = '//*[@id="app"]/div/aside/div/div[2]/div/div/div[1]/div/div[2]/a/div[2]'

clica_em = lambda a: navegador.find_element(By.XPATH, a).click()

def clica_em_campanhas(): 
    try:
        clica_em(campanhas4)
    except NameError as erro:
        clica_em(campanhas2)
    except NameError as erro:
        clica_em(campanhas3)   
    except Exception as  erro:
        clica_em(campanhas1)
        raise    
        
        
   #tentando clicar em 'campanhas' 1 vez a cada 0,5 segundo, 240 vezes, em 2 minutos:
i=0
while (i < 240):
    time.sleep(0.5)
    i += 1
    try:
        clica_em_campanhas()
        print('clicou em campanhas')
    except:
        continue
    else:
        print('sucesso em clicar em campanhas')
        break      


# In[ ]:





# In[ ]:





# In[28]:


#Paths das tabelas:
path_table_1 =  '//*[@id="app"]/div/div[1]/div/div[1]/div[2]/div[2]/div/div[2]/div/div/div[1]/div/div/div[2]'
path_table_2 = '//*[@id="app"]/div/div[1]/div/div[1]/div[2]/div[2]/div/div[2]/div/div/div[2]/div/div'
path_table_5 = '//*[@id="app"]/div/div[1]/div/div[1]/div[2]/div[2]/div/div[2]/div/div/div[5]/div/div'

paths = [path_table_1, path_table_2, path_table_5]

#funão para instanciar elementos:
def instancia_elementos_tables(path):
    return navegador.find_element(By.XPATH, path)


time.sleep(10)

elements = ['element_1', 'element_2', 'element_5']

for i in range(len(paths)):
     elements[i] = instancia_elementos_tables(paths[i])

time.sleep(1)
#Colhe o HTML de todas a tabelas:
lista_html_content = ['html_content_table_1', 'html_content_table_2', 'html_content_table_5']

for i in range(len(paths)):
    lista_html_content[i] = elements[i].get_attribute('outerHTML')
    
    
#transforma o HTML em uma tabela
soup = ['soup_1', 'soup_2', 'soup_5']
table = ['table_1', 'table_2', 'teble_5']

for i in range(len(paths)):
    soup[i] = BeautifulSoup(lista_html_content[i], 'html.parser')
    table[i] = soup[i].find(name='table')
    
df_full_tables = ['df_full_table_1', 'df_full_table_2', 'df_full_table_5']

for i in range(len(paths)):
    df_full_tables[i] =  pd.read_html(str(table[i]))[0]
    
    
#colunas da primeira tabela:
coluna_1_table_1 = df_full_tables[0]['Campanha']['Id']
coluna_2_table_1 = df_full_tables[0]['Campanha']['Nome']
coluna_3_table_1 = df_full_tables[0]['Agentes']['Logados']
coluna_4_table_1 = df_full_tables[0]['Agentes']['Livres']
coluna_5_table_1 = df_full_tables[0]['Agentes']['Trabalhando']
coluna_6_table_1 = df_full_tables[0]['Agentes']['Em Pausa']
coluna_7_table_1 = df_full_tables[0]['Agentes']['Indisp']
coluna_8_table_1 = df_full_tables[0]['Agentes']['Tempo Ocioso']
coluna_9_table_1 = df_full_tables[0]['Mailing']['Mailing']
coluna_10_table_1 = df_full_tables[0]['Total']['Total']
coluna_11_table_1 = df_full_tables[0]['Livres']['Livres']
coluna_12_table_1 = df_full_tables[0]['Sem Contato']['Sem Contato']
coluna_13_table_1 = df_full_tables[0]['Restritos']['Restritos']
coluna_14_table_1 = df_full_tables[0]['Transferidas aos Agentes']['Total']
coluna_15_table_1 = df_full_tables[0]['Transferidas aos Agentes']['Sem Contato']
coluna_16_table_1 = df_full_tables[0]['Transferidas aos Agentes']['Agendamentos']
coluna_17_table_1 = df_full_tables[0]['Transferidas aos Agentes']['Recusas']
coluna_18_table_1 = df_full_tables[0]['Transferidas aos Agentes']['Sucesso']

#separa as colunas da Tabela 2:

coluna_1_table_2 = df_full_tables[1]['Campanha']['Id']
coluna_2_table_2 = df_full_tables[1]['Campanha']['Nome']
coluna_3_table_2 = df_full_tables[1]['Agentes']['Logados']
coluna_4_table_2 = df_full_tables[1]['Agentes']['Livres']
coluna_5_table_2 = df_full_tables[1]['Agentes']['Trabalhando']
coluna_6_table_2 = df_full_tables[1]['Agentes']['Em Pausa']
coluna_7_table_2 = df_full_tables[1]['Nível de Serviço']['Nível de Serviço']
coluna_8_table_2 = df_full_tables[1]['Fila Atual']['Fila Atual']
coluna_9_table_2 = df_full_tables[1]['Tempo Máx Fila']['Tempo Máx Fila']
coluna_10_table_2 = df_full_tables[1]['Chamadas Recebidas']['Chamadas Recebidas']
coluna_11_table_2 = df_full_tables[1]['Transferidas aos Agentes']['Total']
coluna_12_table_2 = df_full_tables[1]['Transferidas aos Agentes']['Sem Contato']
coluna_13_table_2 = df_full_tables[1]['Transferidas aos Agentes']['Agendamentos']
coluna_14_table_2 = df_full_tables[1]['Transferidas aos Agentes']['Recusas']
coluna_15_table_2 = df_full_tables[1]['Transferidas aos Agentes']['Sucesso']
coluna_16_table_2 = df_full_tables[1]['Não Entregue']['Total']
coluna_17_table_2 = df_full_tables[1]['Não Entregue']['Abandonadas']
coluna_18_table_2 = df_full_tables[1]['Não Entregue']['Desviadas']
coluna_19_table_2 = df_full_tables[1]['Não Entregue']['Rejeitadas']

#colunas da 3ª tabela:
coluna_1_table_3 = df_full_tables[2]['Campanha']['Id']
coluna_2_table_3 = df_full_tables[2]['Campanha']['Nome']
coluna_3_table_3 = df_full_tables[2]['Agentes']['Logados']
coluna_4_table_3 = df_full_tables[2]['Agentes']['Livres']
coluna_5_table_3 = df_full_tables[2]['Agentes']['Trabalhando']
coluna_6_table_3 = df_full_tables[2]['Agentes']['Em Pausa']
coluna_7_table_3 = df_full_tables[2]['Agentes']['Indisp']
coluna_8_table_3 = df_full_tables[2]['Agentes']['Tempo Ocioso']
coluna_9_table_3 = df_full_tables[2]['Mailing']['Mailing']
coluna_10_table_3 = df_full_tables[2]['Total']['Total']
coluna_11_table_3 = df_full_tables[2]['Livres']['Livres']
coluna_12_table_3 = df_full_tables[2]['Sem Contato']['Sem Contato']
coluna_13_table_3 = df_full_tables[2]['Transferidas aos Agentes']['Total']
coluna_14_table_3 = df_full_tables[2]['Transferidas aos Agentes']['Sem Contato']
coluna_15_table_3 = df_full_tables[2]['Transferidas aos Agentes']['Agendamentos']
coluna_16_table_3 = df_full_tables[2]['Transferidas aos Agentes']['Recusas']
coluna_17_table_3 = df_full_tables[2]['Transferidas aos Agentes']['Sucesso']


##transformando as colunas 10, 11, 12, 13 e 14 em INT, para qe o excel possa reconhecer seus valores como tal:

#fórmula para tranformar coluna STR em INT:

def transforma_coluna_em_int(coluna):
    
    lista7=[]
    for rows in coluna:
        
        try:
            rows = int(rows)
        except:
            rows = 0
        lista7.append(rows)
    return lista7

#executa a transformação:

coluna_10_table_1 = transforma_coluna_em_int(coluna_10_table_1)
coluna_11_table_1 = transforma_coluna_em_int(coluna_11_table_1)
coluna_12_table_1 = transforma_coluna_em_int(coluna_12_table_1)
coluna_13_table_1 = transforma_coluna_em_int(coluna_13_table_1)
coluna_14_table_1 = transforma_coluna_em_int(coluna_14_table_1)

colunas_table_1 = [
    coluna_1_table_1,
    coluna_2_table_1,
    coluna_3_table_1,
    coluna_4_table_1,
    coluna_5_table_1,
    coluna_6_table_1,
    coluna_7_table_1,
    coluna_8_table_1,
    coluna_9_table_1,
    coluna_10_table_1,
    coluna_11_table_1,
    coluna_12_table_1,
    coluna_13_table_1,
    coluna_14_table_1,
    coluna_15_table_1,
    coluna_16_table_1,
    coluna_17_table_1,
    coluna_18_table_1
]
colunas_table_2 = [
    coluna_1_table_2,
    coluna_2_table_2,
    coluna_3_table_2,
    coluna_4_table_2,
    coluna_5_table_2,
    coluna_6_table_2,
    coluna_7_table_2,
    coluna_8_table_2,
    coluna_9_table_2,
    coluna_10_table_2,
    coluna_11_table_2,
    coluna_12_table_2,
    coluna_13_table_2,
    coluna_14_table_2,
    coluna_15_table_2,
    coluna_16_table_2,
    coluna_17_table_2,
    coluna_18_table_2,
    coluna_19_table_2

]
colunas_table_5 = [
    coluna_1_table_3,
    coluna_2_table_3,
    coluna_3_table_3,
    coluna_4_table_3,
    coluna_5_table_3,
    coluna_6_table_3,
    coluna_7_table_3,
    coluna_8_table_3,
    coluna_9_table_3,
    coluna_10_table_3,
    coluna_11_table_3,
    coluna_12_table_3,
    coluna_13_table_3,
    coluna_14_table_3,
    coluna_15_table_3,
    coluna_16_table_3,
    coluna_17_table_3

]

print(f'colunas_table_1 {len(colunas_table_1)}')
print(f'colunas_table_2 {len(colunas_table_2)}')
print(f'colunas_table_5 {len(colunas_table_5)}')



# ### Abre a planilha modelo:

# In[29]:


#abre a planilha modelo
book = openpyxl.load_workbook('DESEMPENHO_MF_25012022.xlsx') 
#como visualizar sheet's existentes em uma planilha:
print(book.sheetnames)


# ## Coloca os dados da campanha no Modelo:

# In[30]:


painel_page_model_campanha = book['Report_Campaign_015223']


# In[31]:


#função para colar as colunas da CAMPANHA no excel
def transmuta_coluna(coluna, index_column_destino, linha_incial):
    for rows in painel_page_model_campanha.iter_cols(min_col=index_column_destino, max_col=index_column_destino, min_row=(linha_incial + 6)):
        for t in range(663):
            try:
                rows[t].value = coluna[t]
            except:
                continue


# In[32]:


#cola as coluna da tabela CAMPANHA no Modelo Excel:
for i in range(len(colunas_campanha)):
    transmuta_coluna(colunas_campanha[i],(i+1),0)


# ## Coloca os dados da Visao no Modelo:

# In[33]:


painel_page_model_visao = book['Planilha1']


# In[34]:


#função para colar as colunas da VISAO no excel
def transmuta_coluna(coluna, index_column_destino, linha_incial):
    for rows in painel_page_model_visao.iter_cols(min_col=index_column_destino, max_col=index_column_destino, min_row=(linha_incial + 6)):
        for t in range(2500):
            try:
                rows[t].value = coluna[t]
            except:
                continue


# In[35]:


#cola as coluna da tabela VISÃO no Modelo Excel:
for i in range(len(colunas_visao)):
    transmuta_coluna(colunas_visao[i],(i+1),0)


# ## Coloca os dados do Monitor no Modelo:

# In[36]:


#instacia a aba "painel" em uma VA.
painel_page_model = book['Painel']

#função para colar as colunas no excel
def transmuta_coluna(coluna, index_column_destino, linha_incial):
    for rows in painel_page_model.iter_cols(min_col=index_column_destino, max_col=index_column_destino, min_row=(linha_incial + 6)):
        for t in range(663):
            try:
                rows[t].value = coluna[t]
            except:
                continue
                
#cola as coluna da tabela 1 no Modelo Excel
for i in range(len(colunas_table_1)):
    transmuta_coluna(colunas_table_1[i],(i+1),0)
    
#cola as coluna da tabela 2 no Modelo Excel
for j in range(len(colunas_table_2)):
    transmuta_coluna(colunas_table_2[j],(j+1),len(colunas_table_1[1])) 
    
#cola as coluna da tabela 5 no Modelo Excel
for p in range(len(colunas_table_5)):
    transmuta_coluna(colunas_table_5[p],(p+1),(len(colunas_table_1[1]) + len(colunas_table_2[1])))   



# ## colocando a data da atualização:

# In[41]:


atualizacao_page_model = book['atualizacao']

def transmuta_texto(coluna):
    for rows in atualizacao_page_model.iter_cols(min_col=2, max_col=2, min_row=1):
        for t in range(6):
            try:
                rows[t].value = coluna
            except:
                continue
                                                 
transmuta_texto('=AGORA()')


# In[42]:


book.save('teste.xlsx')


# In[43]:


#navegador.close()
navegador.quit()


# In[ ]:





# In[ ]:





# In[ ]:




